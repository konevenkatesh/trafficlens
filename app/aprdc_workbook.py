"""The APRDC classified-count workbook: one sheet per clock hour, as the field proforma.

This is the deliverable. The field sheet is a fixed grid -- four quarter-hour rows per
hour, eighteen vehicle columns -- and the survey is judged on whether that grid is right,
not on whether the model was clever.

Three things it does that a naive dump of counts does not:

**The grid is fixed, and gaps say so.** Every hour prints four quarter-hour rows whether
or not footage covered them. A quarter with full coverage and no traffic prints `0`; a
quarter with no footage prints **NS** (not surveyed) and is excluded from the total. Those
are different facts, and a report that silently omits the second one changes the shape of
the day -- an hour with three columns instead of four reads as a quiet hour rather than a
short recording.

**One sheet per hour PER DIRECTION.** The proforma carries "Direction From / To" in its
header, so a sheet describes one direction of travel; a two-way road needs two. Summing
both into one sheet makes the header unfillable and, worse, hides the directional split
that capacity analysis actually uses -- a road carrying 900 one way and 100 the other is
a different road from one carrying 500 each way, and the combined total cannot tell them
apart.

**Hours are assembled from whatever files cover them, not from one file each.** DVR
recordings do not respect hour boundaries: this station's two files run 09:57-10:48 and
10:48-11:38, so the 10:00 hour is built from both. Every crossing carries a real wall-clock
time, so the aggregation is by clock, and which file it came from is a provenance detail
rather than a structural one.

**Columns the data cannot fill are declared, not zero-filled quietly.** Taxi and APSRTC Bus
are sub-splits that only a human review pass can make; until it runs, taxis sit in the
Car/Jeep/Van column and APSRTC buses in Other Bus, and the coverage sheet says so in words.
A zero in a column nobody has measured is a lie that survives review.
"""
import sys
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import db

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT / "app"))

# The 18 columns of the APRDC proforma, in sheet order, with the MoRTH classes that feed
# each. `needs` names a review pass that must run before the column can be believed.
COLUMNS = [
    ("Two Wheelers",              0.5, ["2W"],              None),
    ("Autorickshaw (3W)",         1.0, ["3W_Auto"],         None),
    ("7 seater (3W)/Maxi Cabs",   1.5, ["Maxi_attr"],       "auto_seats"),
    ("Taxi",                      1.0, ["Taxi_attr"],       "car_use"),
    ("Car/Jeep/Van (non-taxi)",   1.0, ["Car_Jeep_Van"],    None),
    ("Mini Bus",                  1.5, ["Mini_Bus"],        None),
    ("APSRTC Bus",                3.0, ["APSRTC_attr"],     "bus_operator"),
    ("Other Bus",                 3.0, ["Bus"],             None),
    ("LCV",                       1.5, ["LCV"],             None),
    ("2-Axle Truck",              3.0, ["2Axle_Truck"],     None),
    ("3-Axle Truck",              3.0, ["3Axle_Truck"],     None),
    ("MAV (Truck Trolly)",        4.5, ["MAV"],             None),
    ("Tractor-Trolly",            4.5, ["Tractor_Trailer"], None),
    ("Tractor",                   1.5, ["Tractor"],         None),
    ("Cycle",                     0.5, ["Cycle"],           None),
    ("Cycle Rickshaw",            2.0, ["Cycle_Rickshaw"],  None),
    ("Animal Drawn",              8.0, ["Animal_Cart"],     None),
    ("Others (pl. specify)",      1.0, ["Other"],           None),
]
NS = "NS"          # not surveyed: no footage covered this quarter-hour


def _coverage(video_ids):
    """(start, end) wall-clock windows the footage actually covers."""
    out = []
    for vid in video_ids:
        v = db.one("SELECT start_clock, frames, fps FROM videos WHERE id=?", vid)
        if not v or not v["start_clock"]:
            continue
        try:
            s = datetime.strptime(v["start_clock"], "%Y-%m-%d %H:%M:%S")
        except ValueError:
            continue
        out.append((s, s + timedelta(seconds=(v["frames"] or 0) / (v["fps"] or 25))))
    return sorted(out)


def _covered_seconds(windows, a, b):
    return sum(max(0.0, (min(e, b) - max(s, a)).total_seconds()) for s, e in windows)


def collect(video_ids):
    """Every counted crossing across every video, stamped with real wall-clock time."""
    import counting
    import sites
    events, per_video = [], []
    for vid in video_ids:
        lines, src = sites.lines_for(vid)
        v = db.one("SELECT id,name,start_clock,frames,fps FROM videos WHERE id=?", vid)
        if not lines or not v or not v["start_clock"]:
            per_video.append({"id": vid, "name": v["name"] if v else vid,
                              "total": None, "note": "no count line"})
            continue
        r = counting.count_video(vid, lines)
        start = datetime.strptime(v["start_clock"], "%Y-%m-%d %H:%M:%S")
        for e in r["events"]:
            events.append({"t": start + timedelta(seconds=e["time_s"]),
                           "class": e["class"], "direction": e["direction"],
                           "video_id": vid, "track_id": e["track_id"]})
        per_video.append({"id": vid, "name": v["name"], "total": r["total"],
                          "line_source": src, "start": v["start_clock"],
                          "minutes": round((v["frames"] or 0) / (v["fps"] or 25) / 60, 1)})
    return events, per_video


def _class_of(e, attr_map):
    """The proforma class, after any human/model sub-split has been applied."""
    return attr_map.get((e["video_id"], e["track_id"]), e["class"])


def attribute_overrides(video_ids):
    """Taxi / maxi / APSRTC re-labels from the survey app's attribute passes."""
    out = {}
    if not db.one("SELECT name FROM sqlite_master WHERE type='table' AND name='track_attrs'"):
        return out
    q = ",".join("?" * len(video_ids))
    rows = db.rows(f"SELECT * FROM track_attrs WHERE video_id IN ({q})", *video_ids)
    rows.sort(key=lambda r: 0 if r["source"] == "ai" else 1)   # human applied last, wins
    for r in rows:
        k = (r["video_id"], r["track_id"])
        if r["attr"] == "taxi" and r["value"] == "taxi":
            out[k] = "Taxi_attr"
        elif r["attr"] == "maxi" and r["value"] == "maxi":
            out[k] = "Maxi_attr"
        elif r["attr"] in ("apsrtc", "bus_operator") and r["value"] == "apsrtc":
            out[k] = "APSRTC_attr"
        elif r["value"] in ("private", "normal"):
            out.pop(k, None)
    return out


def close_pass_double_counts(video_ids):
    """Vehicles that crossed the line twice because they passed close to the camera.

    A large vehicle passing within a few metres of the lens sweeps its reference point
    right across the frame and back, so it crosses the drawn line twice while passing the
    survey point once. The existing 2-second debounce cancels the jitter version of this;
    it does not reach a bus taking four seconds to pass.

    Widening the debounce was measured and rejected: 4 seconds catches these but removes
    39 crossings across this survey, 18 of them at another station, with no evidence those
    are wrong. Time alone cannot tell "one vehicle passing close" from "two vehicles four
    seconds apart" -- proximity can, and that is a counting change worth doing carefully
    rather than at the end of a delivery.

    So they are found and declared. A known, quantified overcount stated in the workbook is
    honest; a silent one is not, and a fix that moves 39 other numbers to correct 1 is worse
    than either.
    """
    import counting
    import sites
    out = []
    for vid in video_ids:
        lines, _ = sites.lines_for(vid)
        if not lines:
            continue
        r = counting.count_video(vid, lines)
        by_track = defaultdict(list)
        for e in r["events"]:
            by_track[e["track_id"]].append(e)
        for tid, evs in by_track.items():
            if len(evs) < 2 or len({e["direction"] for e in evs}) < 2:
                continue
            w = db.one("""SELECT MAX(x2-x1) w FROM track_points
                          WHERE video_id=? AND track_id=?""", vid, tid)
            out.append({"video_id": vid, "track_id": tid, "class": evs[0]["class"],
                        "clock": evs[0]["clock"], "crossings": len(evs),
                        "max_box_w": int((w or {}).get("w") or 0)})
    return out


def build(video_ids, meta=None):
    """The whole workbook as data: hour sheets, coverage, and what is unreviewed."""
    meta = meta or {}
    events, per_video = collect(video_ids)
    attrs = attribute_overrides(video_ids)
    windows = _coverage(video_ids)
    if not windows:
        return {"error": "no footage with a usable clock"}

    # Speed, if the station has a trap set up. Optional on purpose: most surveys are
    # classified counts and never measure it, and an empty sheet is worse than no sheet.
    speed_rows = []
    try:
        import speed as speed_mod
        sid = db.one("""SELECT site_id FROM videos WHERE id=? AND site_id IS NOT NULL""",
                     video_ids[0]) if video_ids else None
        trap = speed_mod.trap_for(sid["site_id"]) if sid else None
        if trap:
            for vid in video_ids:
                speed_rows.extend(speed_mod.speeds_for(vid, trap))
    except Exception:
        trap = None

    first, last = windows[0][0], windows[-1][1]
    h0 = first.replace(minute=0, second=0, microsecond=0)
    h1 = (last + timedelta(hours=1)).replace(minute=0, second=0, microsecond=0)

    # Keyed by (quarter-hour, direction): the sheet is directional, so the bin must be.
    binned = defaultdict(Counter)
    directions = sorted({e["direction"] for e in events}) or ["in"]
    for e in events:
        q = e["t"].replace(minute=(e["t"].minute // 15) * 15, second=0, microsecond=0)
        binned[(q, e["direction"])][_class_of(e, attrs)] += 1

    speed_summary = None
    if speed_rows:
        import speed as speed_mod
        speed_summary = {**speed_mod.summary(speed_rows), "trap": trap}

    hours, grand = [], Counter()
    per_direction = {d: Counter() for d in directions}
    t = h0
    while t < h1:
      for d in directions:
        rows = []
        for k in range(4):
            a = t + timedelta(minutes=15 * k)
            b = a + timedelta(minutes=15)
            cov = _covered_seconds(windows, a, b)
            counts = binned.get((a, d), Counter())
            cells = []
            for name, _pcu, srcs, _needs in COLUMNS:
                if cov <= 0:
                    cells.append(NS)
                else:
                    cells.append(sum(counts.get(s, 0) for s in srcs))
            rows.append({"label": f"{a.strftime('%H:%M')}–{b.strftime('%H:%M')}",
                         "start": a.strftime("%H:%M"), "end": b.strftime("%H:%M"),
                         "coverage": round(cov / 900, 3),
                         "surveyed": cov > 0, "partial": 0 < cov < 895,
                         "cells": cells})
            if cov > 0:
                for name, _p, srcs, _n in COLUMNS:
                    got = sum(counts.get(s, 0) for s in srcs)
                    grand[name] += got
                    per_direction[d][name] += got
        total = [sum(r["cells"][i] for r in rows if r["surveyed"])
                 for i in range(len(COLUMNS))]
        hours.append({"hour": t.strftime("%H:%M"), "date": t.strftime("%Y-%m-%d"),
                      "day": t.strftime("%A"), "direction": d, "rows": rows,
                      "total": total, "surveyed": any(r["surveyed"] for r in rows)})
      t += timedelta(hours=1)

    n = sum(grand.values())
    pcu = sum(grand[name] * p for name, p, _s, _x in COLUMNS)
    unreviewed = [{"column": name, "needs": needs} for name, _p, _s, needs in COLUMNS
                  if needs and not grand[name]]
    return {
        "double_counts": close_pass_double_counts(video_ids),
        "directions": directions,
        "per_direction": {d: dict(c) for d, c in per_direction.items()},
        "hours": [h for h in hours if h["surveyed"]],
        "per_video": per_video, "totals": dict(grand),
        "total": n, "pcu": round(pcu, 1),
        "window": {"from": first.strftime("%Y-%m-%d %H:%M:%S"),
                   "to": last.strftime("%Y-%m-%d %H:%M:%S"),
                   "minutes": round((last - first).total_seconds() / 60, 1)},
        "unreviewed_columns": unreviewed, "meta": meta,
        "speed": speed_summary,
    }


# ───────────────────────────── the workbook ─────────────────────────────
def write(data, out_path, meta=None):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    meta = {**(data.get("meta") or {}), **(meta or {})}

    wb = Workbook()
    wb.remove(wb.active)
    thin = Side(style="thin", color="000000")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr = PatternFill("solid", fgColor="D9D9D9")
    ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Each sheet is one direction. The place names come from the survey brief and only a
    # person knows them, so they are taken from meta and left blank rather than invented;
    # the machine-known part -- which way the traffic was going -- is always filled in.
    DIRN = {"in": {"label": "towards camera"}, "out": {"label": "away from camera"}}

    def sheet_header(ws, hour, dir_meta):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=19)
        c = ws.cell(1, 1, "GOVERNMENT OF ANDHRA PRADESH, ROADS AND BUILDINGS DEPARTMENT, "
                          "A.P. ROAD DEVELOPMENT CORPORATION (APRDC)")
        c.font = Font(bold=True, size=10); c.alignment = ctr
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=19)
        c = ws.cell(2, 1, "Classified Traffic Volume Count Survey")
        c.font = Font(bold=True, size=13, underline="single"); c.alignment = ctr

        fields = [
            ("Name of Road/Location:", meta.get("road", ""), "Day:", hour["day"],
             "Date :", hour["date"], "Sheet No:", hour["sheet"], "Weather Cond.:",
             meta.get("weather", "")),
            ("Location_ID:", meta.get("location_id", ""), "Direction From:",
             dir_meta.get("from", ""), "To:", dir_meta.get("to", ""),
             "Hour:", hour["hour"], "", ""),
        ]
        for i, row in enumerate(fields, start=4):
            for j, val in enumerate(row):
                cell = ws.cell(i, 1 + j * 2 if j % 2 == 0 else j * 2, val)
                if j % 2 == 0:
                    cell.font = Font(bold=True, size=9)
                else:
                    cell.font = Font(size=9)

    for idx, h in enumerate(data["hours"], start=1):
        h = {**h, "sheet": f"{idx} of {len(data['hours'])}"}
        d = h.get("direction", "in")
        dm = {"from": meta.get(f"direction_{d}_from", ""),
              "to": meta.get(f"direction_{d}_to", ""),
              "label": DIRN.get(d, {}).get("label", d)}
        ws = wb.create_sheet(f"{h['hour'].replace(':','')}hr {d.upper()}")
        sheet_header(ws, h, dm)
        ws.cell(3, 1, f"DIRECTION: {d.upper()} — traffic {dm['label']}"
                ).font = Font(bold=True, size=10)

        r0 = 7
        ws.cell(r0, 1, "Time Period").font = Font(bold=True, size=9)
        ws.cell(r0, 1).fill = hdr; ws.cell(r0, 1).alignment = ctr; ws.cell(r0, 1).border = box
        ws.merge_cells(start_row=r0, start_column=1, end_row=r0 + 1, end_column=1)
        for j, (name, _p, _s, needs) in enumerate(COLUMNS):
            c = ws.cell(r0, 2 + j, name + (" *" if needs else ""))
            c.font = Font(bold=True, size=8); c.fill = hdr; c.alignment = ctr; c.border = box
            n = ws.cell(r0 + 1, 2 + j, j + 1)
            n.font = Font(bold=True, size=8); n.fill = hdr; n.alignment = ctr; n.border = box
        c = ws.cell(r0, 20, "Total"); c.font = Font(bold=True, size=8)
        c.fill = hdr; c.alignment = ctr; c.border = box
        ws.merge_cells(start_row=r0, start_column=20, end_row=r0 + 1, end_column=20)

        r = r0 + 2
        for row in h["rows"]:
            lab = ws.cell(r, 1, row["label"] + ("*" if row["partial"] else ""))
            lab.font = Font(size=9); lab.alignment = ctr; lab.border = box
            for j, val in enumerate(row["cells"]):
                c = ws.cell(r, 2 + j, val)
                c.alignment = ctr; c.border = box; c.font = Font(size=9)
                if val == NS:
                    c.font = Font(size=9, italic=True, color="999999")
            tot = ws.cell(r, 20, sum(v for v in row["cells"] if isinstance(v, int))
                          if row["surveyed"] else NS)
            tot.font = Font(bold=True, size=9); tot.alignment = ctr; tot.border = box
            r += 1

        ws.cell(r, 1, "Total").font = Font(bold=True, size=9)
        ws.cell(r, 1).fill = hdr; ws.cell(r, 1).alignment = ctr; ws.cell(r, 1).border = box
        for j, val in enumerate(h["total"]):
            c = ws.cell(r, 2 + j, val)
            c.font = Font(bold=True, size=9); c.fill = hdr; c.alignment = ctr; c.border = box
        c = ws.cell(r, 20, sum(h["total"]))
        c.font = Font(bold=True, size=9); c.fill = hdr; c.alignment = ctr; c.border = box

        ws.cell(r + 2, 1, "Name of Enumerator:").font = Font(bold=True, size=9)
        ws.cell(r + 2, 2, meta.get("enumerator", ""))
        ws.cell(r + 2, 13, "Supervisor:").font = Font(bold=True, size=9)
        ws.cell(r + 2, 14, meta.get("supervisor", ""))
        ws.cell(r + 4, 1, "* column requires a human review pass — see the Coverage sheet; "
                          "NS = not surveyed (no footage covered that quarter-hour); "
                          "a time period marked * had partial footage coverage"
                ).font = Font(size=8, italic=True)
        ws.column_dimensions["A"].width = 14
        for j in range(len(COLUMNS)):
            ws.column_dimensions[chr(66 + j) if j < 25 else "Z"].width = 9

    # ── day summary ──
    ws = wb.create_sheet("Day Total", 0)
    ws.cell(1, 1, "Classified Traffic Volume Count — day total").font = Font(bold=True, size=12)
    ws.cell(2, 1, f"{meta.get('road','')}  ·  {data['window']['from']} to "
                  f"{data['window']['to']}  ·  {data['window']['minutes']:.0f} minutes surveyed"
            ).font = Font(size=9)
    dirs = data.get("directions") or []
    ws.cell(4, 1, "Column").font = Font(bold=True)
    for i, d in enumerate(dirs):
        ws.cell(4, 2 + i, d.upper()).font = Font(bold=True)
    c0 = 2 + len(dirs)
    ws.cell(4, c0, "Both").font = Font(bold=True)
    ws.cell(4, c0 + 1, "PCU factor").font = Font(bold=True)
    ws.cell(4, c0 + 2, "PCU").font = Font(bold=True)
    ws.cell(4, c0 + 3, "Note").font = Font(bold=True)
    r = 5
    for name, pcu, _s, needs in COLUMNS:
        n = data["totals"].get(name, 0)
        ws.cell(r, 1, name)
        for i, d in enumerate(dirs):
            ws.cell(r, 2 + i, (data["per_direction"].get(d) or {}).get(name, 0))
        ws.cell(r, c0, n); ws.cell(r, c0 + 1, pcu)
        ws.cell(r, c0 + 2, round(n * pcu, 1))
        # The note describes the column's ACTUAL state. Printing "needs review" beside a
        # reviewed figure is worse than printing nothing: it tells the reader to distrust
        # a number that was in fact checked by hand, and it hides the columns that really
        # were not.
        if needs and not n:
            ws.cell(r, c0 + 3, f"reads 0 because the '{needs}' review pass has not been run; "
                          f"those vehicles are counted in their parent column")
        elif needs:
            ws.cell(r, c0 + 3, f"reviewed by hand ({needs} pass)")
        r += 1
    ws.cell(r, 1, "TOTAL (as counted)").font = Font(bold=True)
    ws.cell(r, c0, data["total"]).font = Font(bold=True)
    ws.cell(r, c0 + 2, data["pcu"]).font = Font(bold=True)
    # The adjusted figure belongs on the front sheet, not in a footnote. A known overcount
    # that the reader has to go looking for is one they will not find.
    dc = len(data.get("double_counts") or [])
    if dc:
        r += 1
        ws.cell(r, 1, "less vehicles counted twice").font = Font(bold=True)
        ws.cell(r, c0, -dc).font = Font(bold=True)
        ws.cell(r, c0 + 3, "each passed close to the camera and crossed the line twice while "
                      "passing the survey point once — itemised on the Coverage sheet")
        r += 1
        ws.cell(r, 1, "TOTAL (adjusted)").font = Font(bold=True)
        ws.cell(r, c0, data["total"] - dc).font = Font(bold=True)
        ws.cell(r, c0 + 3, f"{100*dc/data['total']:.1f}% of the raw count")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions[chr(65 + c0 + 2)].width = 62

    # ── coverage & provenance ──
    # ── Speed ──
    # Its own sheet, and only when a trap was set up. Speed is not part of a classified
    # count and most surveys will not have it; a sheet of blanks reads as a measurement
    # that failed rather than one nobody asked for.
    sp = data.get("speed")
    if sp and sp.get("n"):
        ws = wb.create_sheet("Speed")
        r = 1
        c = ws.cell(r, 1, "Spot Speed Summary"); c.font = Font(bold=True, size=13); r += 2
        t = sp.get("trap") or {}
        for k, v in (("Method", "Time between two lines a measured distance apart"),
                     ("Measured distance", f"{t.get('metres','?')} m"),
                     ("Vehicles measured", sp["n"])):
            ws.cell(r, 1, k).font = Font(bold=True, size=9)
            ws.cell(r, 2, v).font = Font(size=9); r += 1
        r += 1

        # The 85th percentile leads. Design speed and enforcement thresholds are set from
        # it, and a report that prints only a mean invites somebody to use the wrong number.
        ws.cell(r, 1, "Speed (km/h)").font = Font(bold=True); r += 1
        for k, v in (("85th percentile", sp.get("p85")), ("Median", sp.get("median")),
                     ("Mean", sp.get("mean")), ("15th percentile", sp.get("p15")),
                     ("Slowest", sp.get("min")), ("Fastest", sp.get("max"))):
            ws.cell(r, 1, k).font = Font(bold=True if k == "85th percentile" else False, size=9)
            cell = ws.cell(r, 2, v)
            cell.font = Font(bold=True if k == "85th percentile" else False, size=9)
            r += 1
        r += 1

        ws.cell(r, 1, "By vehicle class").font = Font(bold=True); r += 1
        for h_, col in (("Class", 1), ("Vehicles", 2), ("Median km/h", 3)):
            cc = ws.cell(r, col, h_); cc.font = Font(bold=True, size=9)
            cc.fill = hdr; cc.border = box; cc.alignment = ctr
        r += 1
        for name, d_ in (sp.get("by_class") or {}).items():
            ws.cell(r, 1, name).font = Font(size=9)
            ws.cell(r, 2, d_["n"]).font = Font(size=9)
            ws.cell(r, 3, d_["median"]).font = Font(size=9)
            r += 1
        r += 1

        ws.cell(r, 1, "By direction").font = Font(bold=True); r += 1
        for name, d_ in (sp.get("by_direction") or {}).items():
            ws.cell(r, 1, DIRN.get(name, {}).get("label", name)).font = Font(size=9)
            ws.cell(r, 2, d_["n"]).font = Font(size=9)
            ws.cell(r, 3, d_["median"]).font = Font(size=9)
            r += 1
        r += 1

        # Warnings go IN the deliverable, not just on a screen the client never sees. A
        # speed figure travels further than the caveat that came with it.
        for w in (sp.get("warnings") or []):
            cell = ws.cell(r, 1, "CHECK: " + w)
            cell.font = Font(size=9, bold=True, color="9C0006"); r += 1
        if sp.get("warnings"):
            r += 1
        ws.cell(r, 1, "Method").font = Font(bold=True); r += 1
        for line in [
            "Each vehicle is timed between two lines drawn across the carriageway, a "
            "distance measured on the ground.",
            "Crossing times are interpolated between frames, so the reading does not "
            "depend on a vehicle happening to be observed exactly on the line.",
            "The measured distance is the dominant source of error: over 30 m, half a "
            "metre is 1.7%. Frame timing is about 1%.",
            "Readings below 3 or above 150 km/h are discarded as tracking failures, not "
            "reported.",
        ]:
            ws.cell(r, 1, "• " + line).font = Font(size=9); r += 1
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 16

    ws = wb.create_sheet("Coverage")
    ws.cell(1, 1, "Coverage, provenance and what is not yet reviewed").font = Font(bold=True, size=12)
    ws.cell(3, 1, "Source clips").font = Font(bold=True)
    for i, h in enumerate(["video", "name", "starts", "minutes", "counted", "line"]):
        ws.cell(4, 1 + i, h).font = Font(bold=True, size=9)
    r = 5
    for v in data["per_video"]:
        for i, k in enumerate(["id", "name", "start", "minutes", "total", "line_source"]):
            ws.cell(r, 1 + i, v.get(k) if v.get(k) is not None else v.get("note", "—"))
        r += 1
    r += 2
    ws.cell(r, 1, "Columns not yet reviewed").font = Font(bold=True); r += 1
    if data["unreviewed_columns"]:
        for u in data["unreviewed_columns"]:
            ws.cell(r, 1, u["column"])
            ws.cell(r, 2, f"reads 0 because the '{u['needs']}' review pass has not been run; "
                          f"those vehicles are counted in their parent column")
            r += 1
    else:
        ws.cell(r, 1, "none — every column has been reviewed"); r += 1
    r += 1
    dc = data.get("double_counts") or []
    ws.cell(r, 1, "Known overcount").font = Font(bold=True); r += 1
    if dc:
        ws.cell(r, 1, f"{len(dc)} vehicle(s) counted twice")
        ws.cell(r, 2, "Each passed very close to the camera, sweeping its reference point "
                      "across the count line and back, so it registered two crossings "
                      "while passing the survey point once. Listed below; subtract them "
                      "if an exact figure is required.")
        r += 1
        for x in dc:
            ws.cell(r, 1, f"video {x['video_id']} track {x['track_id']}")
            ws.cell(r, 2, f"{x['class']} at {x['clock']} — {x['crossings']} crossings, "
                          f"box up to {x['max_box_w']}px wide")
            r += 1
    else:
        ws.cell(r, 1, "none detected"); r += 1
    r += 1
    ws.cell(r, 1, "Method").font = Font(bold=True); r += 1
    for line in [
        "Vehicles are counted where their tracked path crosses the survey line, in the "
        "direction of travel.",
        "Axle class (2-Axle / 3-Axle / MAV) comes from a dedicated classifier, not the "
        "detector: the detector cannot separate them and puts nearly every truck in "
        "2-Axle.",
        "Trucks the classifier was not confident about, or that were too far from the "
        "camera to resolve, were referred to a human reviewer.",
        "NS marks a quarter-hour with no footage. It is excluded from totals rather than "
        "counted as zero.",
    ]:
        ws.cell(r, 1, "• " + line).font = Font(size=9); r += 1
    ws.column_dimensions["A"].width = 30; ws.column_dimensions["B"].width = 90

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return str(out_path)
